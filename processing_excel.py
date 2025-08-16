import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.widgets import RangeSlider, Button
from openpyxl import load_workbook


# Charger la trajectoire de référence une seule fois
REFERENCE_PATH = os.path.join(os.path.dirname(__file__), 'output.xlsx')
try:
    ref_df = pd.read_excel(REFERENCE_PATH)
    ref_x = ref_df.get('x')
    ref_y = ref_df.get('y')
    reference_loaded = ref_x is not None and ref_y is not None
except Exception as e:
    print(f"Erreur lors du chargement de la trajectoire de référence : {e}")
    reference_loaded = False

export_button_created = [False]  # Utiliser une liste pour permettre la mutation dans update()

rayon= 1.0

def is_excel_valid(path):
    try:
        load_workbook(path)
        return True
    except Exception:
        return False


def erreur_cercle(x_arr, y_arr):
    radius = np.sqrt(x_arr**2 + y_arr**2)
    err_abs = np.abs(radius - rayon)
    err_x = x_arr - x_arr * (rayon / radius)
    err_y = y_arr - y_arr * (rayon / radius)
    return err_x, err_y, err_abs

def compute_stats(data, label):
    if len(data) == 0:
        return [label] + ['-']*6
    return [
        label,
        f"{np.mean(data):.4f}",
        f"{np.var(data):.4f}",
        f"{np.std(data):.4f}",
        f"{np.min(data):.4f}",
        f"{np.max(data):.4f}",
        f"{np.median(data):.4f}",
    ]
    
def quaternion_to_yaw(x, y, z, w):
    # yaw (rotation autour de Z)
    siny_cosp = 2 * (w * z + x * y)
    cosy_cosp = 1 - 2 * (y**2 + z**2)
    return np.arctan2(siny_cosp, cosy_cosp)

def compute_ref_orientation(ref_x, ref_y):
    dx = np.diff(ref_x)
    dy = np.diff(ref_y)
    return np.arctan2(dy, dx)

def orientation_towards_center(x, y, center=(0, 0)):
    dx = center[0] - x
    dy = center[1] - y
    return np.arctan2(dy, dx)

def orientation_tangent_to_circle(x, y, center=(0, 0)):
    # Orientation vers le centre
    dx = center[0] - x
    dy = center[1] - y
    radial_angle = np.arctan2(dy, dx)
    # Orientation tangente = orientation radiale + π/2
    return radial_angle + np.pi/2

def angle_difference(a1, a2):
    diff = a1 - a2
    return (diff + np.pi) % (2 * np.pi) - np.pi

def compute_velocities(timestamps, positions, orientations):
    """
    Calcule les vitesses linéaires et angulaires à partir des positions et orientations.
    
    Args:
        timestamps: array des temps
        positions: array des positions (x, y) shape (n, 2)
        orientations: array des orientations yaw (n,)
    
    Returns:
        linear_velocities: vitesses linéaires (vx, vy, v_norm) shape (n-1, 3)
        angular_velocities: vitesses angulaires (n-1,)
        timestamps_vel: timestamps pour les vitesses (n-1,)
    """
    if len(timestamps) < 2:
        return np.array([]), np.array([]), np.array([])
    
    # Calcul des vitesses linéaires
    dt = np.diff(timestamps)
    dx = np.diff(positions[:, 0])
    dy = np.diff(positions[:, 1])
    
    vx = dx / dt
    vy = dy / dt
    v_norm = np.sqrt(vx**2 + vy**2)
    
    linear_velocities = np.column_stack((vx, vy, v_norm))
    
    # Calcul des vitesses angulaires
    dyaw = np.diff(orientations)
    # Correction pour les discontinuités de ±π
    dyaw = (dyaw + np.pi) % (2*np.pi) - np.pi
    angular_velocities = dyaw / dt
    
    # Timestamps pour les vitesses (moyennes des temps adjacents)
    timestamps_vel = timestamps[:-1] + dt/2
    
    return linear_velocities, angular_velocities, timestamps_vel

def calculer_erreurs_stats(df_sync, ref_x, ref_y, t_start, t_end):
    """Calcule toutes les erreurs et statistiques sur la plage temporelle donnée."""
    df_sync_f = df_sync[(df_sync['timestamp'] >= t_start) & (df_sync['timestamp'] <= t_end)]
    mask_uav = df_sync_f['prefix'] == '/vrpn_client_node/Drone_0/pose'
    mask_ugv = df_sync_f['prefix'] == '/vrpn_client_node/ugv/pose'

    uav_x = df_sync_f.loc[mask_uav, 'pos_x'].to_numpy()
    uav_y = -df_sync_f.loc[mask_uav, 'pos_z'].to_numpy()
    ugv_x = df_sync_f.loc[mask_ugv, 'pos_x'].to_numpy()
    ugv_y = -df_sync_f.loc[mask_ugv, 'pos_z'].to_numpy()

    # Erreurs UAV/référence
    ref_points = np.column_stack((np.array(ref_x), np.array(ref_y)))
    err_uav_x, err_uav_y, err_uav_abs = [], [], []
    for x, y in zip(uav_x, uav_y):
        dists = np.sqrt((ref_points[:,0] - x)**2 + (ref_points[:,1] - y)**2)
        idx_min = np.argmin(dists)
        closest_x, closest_y = ref_points[idx_min]
        err_uav_x.append(x - closest_x)
        err_uav_y.append(y - closest_y)
        err_uav_abs.append(dists[idx_min])

    # Erreurs UGV/UAV
    min_len = min(len(uav_x), len(ugv_x))
    err_ugv_x = ugv_x[:min_len] - uav_x[:min_len]
    err_ugv_y = ugv_y[:min_len] - uav_y[:min_len]
    err_ugv_abs = np.sqrt(err_ugv_x**2 + err_ugv_y**2)

    # Erreurs UGV/référence
    err_ugv_x_ref, err_ugv_y_ref, err_ugv_abs_ref = [], [], []
    for x, y in zip(ugv_x, ugv_y):
        dists = np.sqrt((ref_points[:,0] - x)**2 + (ref_points[:,1] - y)**2)
        idx_min = np.argmin(dists)
        closest_x, closest_y = ref_points[idx_min]
        err_ugv_x_ref.append(x - closest_x)
        err_ugv_y_ref.append(y - closest_y)
        err_ugv_abs_ref.append(dists[idx_min])

    uav_err_circ_x, uav_err_circ_y, uav_err_circ_abs = erreur_cercle(uav_x, uav_y)
    ugv_err_circ_x, ugv_err_circ_y, ugv_err_circ_abs = erreur_cercle(ugv_x, ugv_y)
    
    # Récupération des quaternions
    ori_cols = ['ori_x', 'ori_y', 'ori_z', 'ori_w']
    uav_ori = df_sync_f.loc[mask_uav, ori_cols].to_numpy()
    ugv_ori = df_sync_f.loc[mask_ugv, ori_cols].to_numpy()

    # Convertir en yaw
    uav_yaws = np.array([quaternion_to_yaw(*q) for q in uav_ori])
    ugv_yaws = np.array([quaternion_to_yaw(*q) for q in ugv_ori])

    # Référence orientation (trajectoire pour UAV, cercle pour UAV/UGV)
    # Pour UAV : tangente à la trajectoire (calculée via les ref_x, ref_y)
    ref_orientations = compute_ref_orientation(np.array(ref_x), np.array(ref_y))

    # Interpolation temporelle de l'orientation de référence à la position UAV
    from scipy.spatial import cKDTree
    ref_pts = np.column_stack((ref_x[:-1], ref_y[:-1]))
    tree = cKDTree(ref_pts)
    uav_pts = np.column_stack((uav_x, uav_y))
    _, idxs = tree.query(uav_pts)
    uav_ref_yaw = ref_orientations[idxs]
    
    ugv_pts = np.column_stack((ugv_x, ugv_y))
    _, idxs = tree.query(ugv_pts)
    ugv_ref_yaw = ref_orientations[idxs]


    # Pour le cercle : orientation vers le centre (0,0)
    uav_circle_yaw = orientation_towards_center(uav_x, uav_y)
    ugv_circle_yaw = orientation_tangent_to_circle(ugv_x, ugv_y)

    # Calcul des erreurs d'orientation
    uav_ori_err_traj = angle_difference(uav_yaws, uav_ref_yaw)
    ugv_ori_err_traj = angle_difference(ugv_yaws, ugv_ref_yaw)
    uav_ori_err_circ = angle_difference(uav_yaws, uav_circle_yaw)
    ugv_ori_err_circ = angle_difference(ugv_yaws, ugv_circle_yaw)
    ori_diff_uav_ugv = angle_difference(uav_yaws[:min_len], ugv_yaws[:min_len])

    return [
        compute_stats(err_uav_x, "Erreur X UAV/Référence"),
        compute_stats(err_uav_y, "Erreur Y UAV/Référence"),
        compute_stats(err_uav_abs, "Erreur abs UAV/Référence"),
        compute_stats(err_ugv_x_ref, "Erreur X UGV/Référence"),
        compute_stats(err_ugv_y_ref, "Erreur Y UGV/Référence"),
        compute_stats(err_ugv_abs_ref, "Erreur abs UGV/Référence"),
        compute_stats(err_ugv_x, "Erreur X UGV/UAV"),
        compute_stats(err_ugv_y, "Erreur Y UGV/UAV"),
        compute_stats(err_ugv_abs, "Erreur abs UGV/UAV"),
        compute_stats(uav_err_circ_x, "Erreur X UAV/Cercle"),
        compute_stats(uav_err_circ_y, "Erreur Y UAV/Cercle"),
        compute_stats(ugv_err_circ_x, "Erreur X UGV/Cercle"),
        compute_stats(ugv_err_circ_y, "Erreur Y UGV/Cercle"),
        compute_stats(uav_err_circ_abs, "Erreur abs UAV/Cercle"),
        compute_stats(ugv_err_circ_abs, "Erreur abs UGV/Cercle"),
        compute_stats(uav_ori_err_traj, "Erreur orientation UAV Trajectoire"),
        compute_stats(ugv_ori_err_traj, "Erreur orientation UGV Trajectoire"),
        compute_stats(uav_ori_err_circ, "Erreur orientation UAV Cercle"),
        compute_stats(ugv_ori_err_circ, "Erreur orientation UGV Cercle"),
        compute_stats(ori_diff_uav_ugv, "Différence orientation UAV/UGV"),
    ]


# --- Nouvelle fonction pour calculer et afficher les erreurs ---
def plot_erreurs(fichier_excel, ref_x, ref_y):

    # Lecture du DataFrame et calcul des bornes de timestamp
    try:
        df_sync = pd.read_excel(fichier_excel, sheet_name='poses_sync')
    except Exception as e:
        print(f"Erreur lecture poses_sync : {e}")
        return
    timestamps = df_sync['timestamp'].to_numpy()
    t_min, t_max = np.min(timestamps), np.max(timestamps)

    # --- Fenêtre 1 : erreurs par rapport à la référence ---
    fig1, axs1 = plt.subplots(3, 2, figsize=(12, 10))  # augmente la taille
    maximize_figure()

    # plt.subplots_adjust(bottom=0.22)
    l_uav_x, = axs1[0, 0].plot([], label='Erreur x UAV/référence')
    l_uav_y, = axs1[0, 1].plot([], label='Erreur y UAV/référence')
    l_uav_abs, = axs1[1, 0].plot([], label='Erreur absolue UAV/référence')
    l_ugv_x, = axs1[1, 1].plot([], label='Erreur x UGV/UAV', color='green')
    l_ugv_y, = axs1[2, 0].plot([], label='Erreur y UGV/UAV', color='green')
    l_ugv_abs, = axs1[2, 1].plot([], label='Erreur absolue UGV/UAV', color='green')
    for ax in axs1.flat:
        ax.legend(); ax.set_xlabel('Temps (s)'); ax.set_ylabel('Erreur (m)'); ax.grid(True)
    fig1.suptitle('Erreurs par rapport à la trajectoire de référence')
    # Boutons save
    save_buttons1 = []
    for i, ax in enumerate(axs1.flat):
        ax_save = plt.axes([0.01 + 0.16*i, 0.01, 0.15, 0.05], figure=fig1)
        btn = Button(ax_save, f'Save {ax.get_title() or ax.get_legend().get_texts()[0].get_text()}')
        save_buttons1.append(btn)

    # --- Fenêtre 2 : erreurs cercle et trajectoires ---
    fig2, axs2 = plt.subplots(2, 2, figsize=(12, 8))
    maximize_figure()

    # plt.subplots_adjust(bottom=0.22)
    l_ugv_x_ref, = axs2[0, 0].plot([], label='Erreur x UGV/référence', color='orange')
    l_ugv_y_ref, = axs2[0, 1].plot([], label='Erreur y UGV/référence', color='orange')
    l_ugv_abs_ref, = axs2[1, 0].plot([], label='Erreur absolue UGV/référence', color='orange')
    ax_traj = axs2[1, 1]
    l_traj_drone, = ax_traj.plot([], [], label='Trajectoire Drone')
    l_traj_ugv, = ax_traj.plot([], [], label='Trajectoire UGV', color='green')
    l_traj_ref, = ax_traj.plot([], [], color='red', linestyle='--', label='Trajectoire de référence') if ref_x is not None and ref_y is not None else (None,)
    theta = np.linspace(0, 2*np.pi, 200)
    circle_x = rayon * np.cos(theta)
    circle_y = rayon * np.sin(theta)
    l_traj_cercle, = ax_traj.plot(circle_x, circle_y, color='blue', linestyle=':', label=f'Cercle de rayon {rayon}m')
    ax_traj.set_xlabel('x (m)')
    ax_traj.set_ylabel('y (m)')
    ax_traj.axis('equal')
    ax_traj.grid(True)
    ax_traj.legend()
    for ax in axs2.flat:
        ax.legend(); ax.set_xlabel('Temps (s)'); ax.set_ylabel('Erreur (m)'); ax.grid(True)
    fig2.suptitle('Erreurs UGV/référence et trajectoires')
    # Boutons save
    save_buttons2 = []
    for i, ax in enumerate(axs2.flat):
        ax_save = plt.axes([0.15 + 0.2*i, 0.01, 0.15, 0.05], figure=fig2)
        btn = Button(ax_save, f'Save {ax.get_title() or ax.get_legend().get_texts()[0].get_text()}')
        save_buttons2.append(btn)

    # --- Fenêtre 3 : slider indépendant ---
    fig_slider = plt.figure(figsize=(8, 2))
    fig_slider.suptitle('Contrôle du timestamp')
    ax_slider = fig_slider.add_axes([0.15, 0.4, 0.7, 0.2])
    slider = RangeSlider(ax_slider, 'Timestamp', t_min, t_max, valinit=(t_min, t_max))
    
    # --- Fenêtre 4 : statistiques des erreurs ---
    fig_stats, axs_stats = plt.subplots(figsize=(60, 6))
    maximize_figure()
    fig_stats.suptitle('Statistiques des erreurs (plage sélectionnée)')
    table_stats_holder = [None]  # list to hold table object
    empty_data = [[""] * 7]
    table = axs_stats.table(
        cellText=empty_data,
        colLabels=["Erreur", "Moyenne", "Variance", "Écart-type", "Min", "Max", "Médiane"],
        loc='center'
    )
    table.scale(1, 2)
    axs_stats.axis('off')
    table_stats_holder[0] = table
    axs_stats.axis('off')  # Cache les axes
    
    # --- Fenêtre 5 : erreurs par rapport au cercle ---
    fig5, axs5 = plt.subplots(3, 2, figsize=(12, 8))
    fig5.suptitle(f"Erreurs par rapport à un cercle de rayon {rayon}m (centre 0,0)")
    maximize_figure()
    l_uav_err_circ_x, = axs5[0, 0].plot([], label='Erreur X UAV/cercle')
    l_uav_err_circ_y, = axs5[0, 1].plot([], label='Erreur Y UAV/cercle')
    l_ugv_err_circ_x, = axs5[1, 0].plot([], label='Erreur X UGV/cercle')
    l_ugv_err_circ_y, = axs5[1, 1].plot([], label='Erreur Y UGV/cercle')
    l_ugv_err_circ_abs, = axs5[2, 0].plot([], label='Erreur absolue UGV/cercle')
    l_uav_err_circ_abs, = axs5[2, 1].plot([], label='Erreur absolue UAV/cercle')
    for ax in axs5.flat:
        ax.legend(); ax.set_xlabel('Temps (s)'); ax.set_ylabel('Erreur (m)'); ax.grid(True)
       
    # --- Boutons save pour la fenêtre 5 ---
    save_buttons5 = []
    for i, ax in enumerate(axs5.flat):
        ax_save = plt.axes([0.01 + 0.16*i, 0.01, 0.15, 0.05], figure=fig5)
        btn = Button(ax_save, f'Save {ax.get_title() or ax.get_legend().get_texts()[0].get_text()}')
        save_buttons5.append(btn)
    
    # --- Fenêtre 6 : erreurs d'orientation ---
    fig6, axs6 = plt.subplots(2, 3, figsize=(12, 8))
    fig6.suptitle("Erreurs d'orientation (Yaw)")
    l_uav_ori_err_traj, = axs6[0, 0].plot([], [], label='Erreur orientation UAV trajectoire')
    l_ugv_ori_err_taj, = axs6[0, 1].plot([], [], label='Erreur orientation UGV trajectoire')
    l_uav_ori_err_cir, = axs6[1, 0].plot([], [], label='Erreur orientation UAV cercle')
    l_ugv_ori_err_cir, = axs6[1, 1].plot([], [], label='Erreur orientation UGV cercle')
    l_ori_diff_uav_ugv, = axs6[0, 2].plot([], [], label='Erreur orientation Drone_0 vs UGV', color='purple')

    for ax in axs6.flat:
        ax.set_xlabel('Temps (s)')
        ax.set_ylabel("Erreur d'angle (rad)")
        ax.grid(True)
        ax.legend()

    maximize_figure()
    
    save_buttons6 = []
    for i, ax in enumerate(axs6.flat):
        ax_save = plt.axes([0.01 + 0.24*i, 0.01, 0.2, 0.05], figure=fig6)
        btn = Button(ax_save, f'Save {i}')
        save_buttons6.append(btn)

    for i, btn in enumerate(save_buttons6):
        btn.on_clicked(lambda event, ax=axs6.flat[i], fig=fig6, name=f'graph6_{i}': save_graph(ax, fig, name))
    
    # --- NOUVELLE Fenêtre 7 : Vitesses UAV et UGV ---
    fig7, axs7 = plt.subplots(2, 3, figsize=(15, 10))
    fig7.suptitle("Vitesses UAV et UGV")
    maximize_figure()
    
    # Vitesses linéaires UAV
    l_uav_vx, = axs7[0, 0].plot([], [], label='Vitesse X UAV', color='blue')
    l_uav_vy, = axs7[0, 1].plot([], [], label='Vitesse Y UAV', color='blue')
    l_uav_vnorm, = axs7[0, 2].plot([], [], label='Vitesse norme UAV', color='blue')
    
    # Vitesses linéaires UGV
    l_ugv_vx, = axs7[1, 0].plot([], [], label='Vitesse X UGV', color='green')
    l_ugv_vy, = axs7[1, 1].plot([], [], label='Vitesse Y UGV', color='green')
    l_ugv_vnorm, = axs7[1, 2].plot([], [], label='Vitesse norme UGV', color='green')
    
    # Configuration des axes
    axs7[0, 0].set_title('Vitesses X')
    axs7[0, 1].set_title('Vitesses Y') 
    axs7[0, 2].set_title('Vitesses normalisées')
    
    for ax in axs7.flat:
        ax.set_xlabel('Temps (s)')
        ax.set_ylabel('Vitesse (m/s)')
        ax.grid(True)
        ax.legend()
    
    # Boutons save pour la fenêtre 7
    save_buttons7 = []
    for i, ax in enumerate(axs7.flat):
        ax_save = plt.axes([0.01 + 0.16*i, 0.01, 0.15, 0.05], figure=fig7)
        btn = Button(ax_save, f'Save Vel {i}')
        save_buttons7.append(btn)
    
    # --- NOUVELLE Fenêtre 8 : Vitesses angulaires UAV et UGV ---
    fig8, axs8 = plt.subplots(1, 2, figsize=(12, 6))
    fig8.suptitle("Vitesses angulaires UAV et UGV")
    maximize_figure()
    
    l_uav_omega, = axs8[0].plot([], [], label='Vitesse angulaire UAV', color='blue')
    l_ugv_omega, = axs8[1].plot([], [], label='Vitesse angulaire UGV', color='green')
    
    axs8[0].set_title('Vitesse angulaire UAV')
    axs8[1].set_title('Vitesse angulaire UGV')
    
    for ax in axs8:
        ax.set_xlabel('Temps (s)')
        ax.set_ylabel('Vitesse angulaire (rad/s)')
        ax.grid(True)
        ax.legend()
    
    # Boutons save pour la fenêtre 8
    save_buttons8 = []
    for i, ax in enumerate(axs8):
        ax_save = plt.axes([0.2 + 0.3*i, 0.01, 0.25, 0.05], figure=fig8)
        btn = Button(ax_save, f'Save AngVel {i}')
        save_buttons8.append(btn)

    def export_all_stats():
        output_excel = os.path.join(dossier_racine, 'stats_globales.xlsx')
        writer = pd.ExcelWriter(output_excel, engine='openpyxl')

        # On prend la plage actuelle du slider
        t_start, t_end = slider.val

        for folder_name in os.listdir(dossier_racine):
            folder_path = os.path.join(dossier_racine, folder_name)
            if not os.path.isdir(folder_path):
                continue
            fichiers = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
            if not fichiers:
                continue
            fichier_excel_path = os.path.join(folder_path, fichiers[0])
            try:
                df_sync_local = pd.read_excel(fichier_excel_path, sheet_name='poses_sync')
            except Exception:
                continue

            stats_data = calculer_erreurs_stats(df_sync_local, ref_x, ref_y, t_start, t_end)
            df_stats = pd.DataFrame(stats_data, columns=["Erreur", "Moyenne", "Variance", "Écart-type", "Min", "Max", "Médiane"])
            df_stats.to_excel(writer, sheet_name=folder_name[:31], index=False)

        writer.close()
        print(f"Statistiques exportées dans : {output_excel}")

    # --- Update function commune ---
    def update(val):
        t_start, t_end = slider.val
        # Filtrage sur le DataFrame
        df_sync_f = df_sync[(df_sync['timestamp'] >= t_start) & (df_sync['timestamp'] <= t_end)]
        # UAV
        mask_uav_f = df_sync_f['prefix'] == '/vrpn_client_node/Drone_0/pose'
        uav_x_f = df_sync_f.loc[mask_uav_f, 'pos_x'].to_numpy()
        uav_y_f = -df_sync_f.loc[mask_uav_f, 'pos_z'].to_numpy()
        # UGV
        mask_ugv_f = df_sync_f['prefix'] == '/vrpn_client_node/ugv/pose'
        ugv_x_f = df_sync_f.loc[mask_ugv_f, 'pos_x'].to_numpy()
        ugv_y_f = -df_sync_f.loc[mask_ugv_f, 'pos_z'].to_numpy()
        # UAV timestamps
        timestamps_uav = df_sync_f.loc[mask_uav_f, 'timestamp'].to_numpy() - t_start
        # UGV timestamps
        timestamps_ugv = df_sync_f.loc[mask_ugv_f, 'timestamp'].to_numpy() - t_start
        
        # Calcul des vitesses
        # UAV
        uav_positions = np.column_stack((uav_x_f, uav_y_f))
        ori_cols = ['ori_x', 'ori_y', 'ori_z', 'ori_w']
        uav_ori = df_sync_f.loc[mask_uav_f, ori_cols].to_numpy()
        uav_yaws = np.array([quaternion_to_yaw(*q) for q in uav_ori])
        uav_linear_vel, uav_angular_vel, uav_timestamps_vel = compute_velocities(
            timestamps_uav + t_start, uav_positions, uav_yaws
        )
        uav_timestamps_vel_rel = uav_timestamps_vel - t_start
        
        # UGV
        ugv_positions = np.column_stack((ugv_x_f, ugv_y_f))
        ugv_ori = df_sync_f.loc[mask_ugv_f, ori_cols].to_numpy()
        ugv_yaws = np.array([quaternion_to_yaw(*q) for q in ugv_ori])
        ugv_linear_vel, ugv_angular_vel, ugv_timestamps_vel = compute_velocities(
            timestamps_ugv + t_start, ugv_positions, ugv_yaws
        )
        ugv_timestamps_vel_rel = ugv_timestamps_vel - t_start
        
        # --- Calcul stats vitesses UAV ---
        stats_vitesses = [
            compute_stats(uav_linear_vel[:,0] if len(uav_linear_vel) else [], "Vitesse X UAV"),
            compute_stats(uav_linear_vel[:,1] if len(uav_linear_vel) else [], "Vitesse Y UAV"),
            compute_stats(uav_linear_vel[:,2] if len(uav_linear_vel) else [], "Vitesse Norme UAV"),
            compute_stats(uav_angular_vel if len(uav_angular_vel) else [], "Vitesse angulaire UAV"),
            compute_stats(ugv_linear_vel[:,0] if len(ugv_linear_vel) else [], "Vitesse X UGV"),
            compute_stats(ugv_linear_vel[:,1] if len(ugv_linear_vel) else [], "Vitesse Y UGV"),
            compute_stats(ugv_linear_vel[:,2] if len(ugv_linear_vel) else [], "Vitesse Norme UGV"),
            compute_stats(ugv_angular_vel if len(ugv_angular_vel) else [], "Vitesse angulaire UGV"),
        ]

        
        # Erreur UAV/référence
        err_uav_x_f, err_uav_y_f, err_uav_abs_f = [], [], []
        ref_points = np.column_stack((np.array(ref_x), np.array(ref_y)))
        for x, y in zip(uav_x_f, uav_y_f):
            dists = np.sqrt((ref_points[:,0] - x)**2 + (ref_points[:,1] - y)**2)
            idx_min = np.argmin(dists)
            closest_x, closest_y = ref_points[idx_min]
            err_uav_x_f.append(x - closest_x)
            err_uav_y_f.append(y - closest_y)
            err_uav_abs_f.append(dists[idx_min])
        err_uav_x_f = np.array(err_uav_x_f)
        err_uav_y_f = np.array(err_uav_y_f)
        err_uav_abs_f = np.array(err_uav_abs_f)
        # Erreur UGV/UAV
        min_len = min(len(uav_x_f), len(ugv_x_f))
        err_ugv_x_f = ugv_x_f[:min_len] - uav_x_f[:min_len]
        err_ugv_y_f = ugv_y_f[:min_len] - uav_y_f[:min_len]
        err_ugv_abs_f = np.sqrt(err_ugv_x_f**2 + err_ugv_y_f**2)
        # Timestamps synchronisés pour UGV/UAV
        timestamps_sync = timestamps_uav[:min_len]

        # Erreur UGV/référence
        err_ugv_x_ref_f, err_ugv_y_ref_f, err_ugv_abs_ref_f = [], [], []
        for x, y in zip(ugv_x_f, ugv_y_f):
            dists = np.sqrt((ref_points[:,0] - x)**2 + (ref_points[:,1] - y)**2)
            idx_min = np.argmin(dists)
            closest_x, closest_y = ref_points[idx_min]
            err_ugv_x_ref_f.append(x - closest_x)
            err_ugv_y_ref_f.append(y - closest_y)
            err_ugv_abs_ref_f.append(dists[idx_min])
        err_ugv_x_ref_f = np.array(err_ugv_x_ref_f)
        err_ugv_y_ref_f = np.array(err_ugv_y_ref_f)
        err_ugv_abs_ref_f = np.array(err_ugv_abs_ref_f)

        # Update courbes fenêtre 1
        l_uav_x.set_data(timestamps_uav, err_uav_x_f)
        l_uav_y.set_data(timestamps_uav, err_uav_y_f)
        l_uav_abs.set_data(timestamps_uav, err_uav_abs_f)
        l_ugv_abs.set_data(timestamps_sync, err_ugv_abs_f)
        l_ugv_x.set_data(timestamps_sync, err_ugv_x_f)
        l_ugv_y.set_data(timestamps_sync, err_ugv_y_f)

        # Update courbes vitesses fenêtre 7
        if len(uav_linear_vel) > 0:
            l_uav_vx.set_data(uav_timestamps_vel_rel, uav_linear_vel[:, 0])
            l_uav_vy.set_data(uav_timestamps_vel_rel, uav_linear_vel[:, 1])
            l_uav_vnorm.set_data(uav_timestamps_vel_rel, uav_linear_vel[:, 2])
        else:
            l_uav_vx.set_data([], [])
            l_uav_vy.set_data([], [])
            l_uav_vnorm.set_data([], [])
            
        if len(ugv_linear_vel) > 0:
            l_ugv_vx.set_data(ugv_timestamps_vel_rel, ugv_linear_vel[:, 0])
            l_ugv_vy.set_data(ugv_timestamps_vel_rel, ugv_linear_vel[:, 1])
            l_ugv_vnorm.set_data(ugv_timestamps_vel_rel, ugv_linear_vel[:, 2])
        else:
            l_ugv_vx.set_data([], [])
            l_ugv_vy.set_data([], [])
            l_ugv_vnorm.set_data([], [])

        # Update courbes vitesses angulaires fenêtre 8
        if len(uav_angular_vel) > 0:
            l_uav_omega.set_data(uav_timestamps_vel_rel, uav_angular_vel)
        else:
            l_uav_omega.set_data([], [])
            
        if len(ugv_angular_vel) > 0:
            l_ugv_omega.set_data(ugv_timestamps_vel_rel, ugv_angular_vel)
        else:
            l_ugv_omega.set_data([], [])

        # Limites dynamiques pour erreurs (fenêtre 1)
        for i, ax in enumerate(axs1.flat):
            # Récupère les données y du plot
            lines = ax.get_lines()
            y_arrays = [line.get_ydata() for line in lines if len(line.get_ydata()) > 0]
            if len(y_arrays) > 0:
                ydata = np.concatenate(y_arrays)
                lim = max(1, np.ceil(np.max(np.abs(ydata))))
                ax.set_ylim(-lim, lim)
            else:
                ax.set_ylim(-1, 1)
            ax.relim(); ax.autoscale_view()

        # Update courbes fenêtre 2
        l_ugv_x_ref.set_data(timestamps_ugv, err_ugv_x_ref_f)
        l_ugv_y_ref.set_data(timestamps_ugv, err_ugv_y_ref_f)
        l_ugv_abs_ref.set_data(timestamps_ugv, err_ugv_abs_ref_f)
        l_traj_drone.set_data(uav_x_f, uav_y_f)
        l_traj_ugv.set_data(ugv_x_f, ugv_y_f)
        if l_traj_ref:
            l_traj_ref.set_data(ref_x, ref_y)

        # Limites dynamiques pour erreurs (fenêtre 2, sauf trajectoire)
        for i, ax in enumerate(axs2.flat):
            if ax == ax_traj:
                ax.set_xlabel('x (m)')
                ax.set_ylabel("y (m)")
                continue
            lines = ax.get_lines()
            y_arrays = [line.get_ydata() for line in lines if len(line.get_ydata()) > 0]
            if len(y_arrays) > 0:
                ydata = np.concatenate(y_arrays)
                lim = max(1, np.ceil(np.max(np.abs(ydata))))
                ax.set_ylim(-lim, lim)
            else:
                ax.set_ylim(-1, 1)
            ax.relim(); ax.autoscale_view()

        # Limites dynamiques pour la trajectoire
        x_arrays = [l_traj_drone.get_xdata(), l_traj_ugv.get_xdata()]
        y_arrays = [l_traj_drone.get_ydata(), l_traj_ugv.get_ydata()]
        if l_traj_ref:
            x_arrays.append(np.array(ref_x))
            y_arrays.append(np.array(ref_y))
        x_arrays = [arr for arr in x_arrays if len(arr) > 0]
        y_arrays = [arr for arr in y_arrays if len(arr) > 0]
        if len(x_arrays) > 0:
            xdata = np.concatenate(x_arrays)
            lim_x = max(1, np.ceil(np.max(np.abs(xdata))))
        else:
            lim_x = 1
        if len(y_arrays) > 0:
            ydata = np.concatenate(y_arrays)
            lim_y = max(1, np.ceil(np.max(np.abs(ydata))))
        else:
            lim_y = 1
        ax_traj.set_xlim(-lim_x, lim_x)
        ax_traj.set_ylim(-lim_y, lim_y)
        ax_traj.relim(); ax_traj.autoscale_view()

        # Limites dynamiques pour vitesses fenêtre 7
        for ax in axs7.flat:
            lines = ax.get_lines()
            y_arrays = [line.get_ydata() for line in lines if len(line.get_ydata()) > 0]
            if len(y_arrays) > 0:
                ydata = np.concatenate(y_arrays)
                if len(ydata) > 0:
                    y_min, y_max = np.min(ydata), np.max(ydata)
                    margin = max(0.1, (y_max - y_min) * 0.1)
                    ax.set_ylim(y_min - margin, y_max + margin)
            ax.relim(); ax.autoscale_view()

        # Limites dynamiques pour vitesses angulaires fenêtre 8
        for ax in axs8:
            lines = ax.get_lines()
            y_arrays = [line.get_ydata() for line in lines if len(line.get_ydata()) > 0]
            if len(y_arrays) > 0:
                ydata = np.concatenate(y_arrays)
                if len(ydata) > 0:
                    y_min, y_max = np.min(ydata), np.max(ydata)
                    margin = max(0.1, (y_max - y_min) * 0.1)
                    ax.set_ylim(y_min - margin, y_max + margin)
            ax.relim(); ax.autoscale_view()

        fig1.canvas.draw_idle()
        fig2.canvas.draw_idle()
        fig7.canvas.draw_idle()
        fig8.canvas.draw_idle()
            
        # --- Erreur par rapport au cercle de rayon 1m centré en (0, 0) ---
        def erreur_cercle(x_arr, y_arr):
            radius = np.sqrt(x_arr**2 + y_arr**2)
            err_abs = np.abs(radius - rayon)  
            err_x = x_arr - x_arr * (rayon / radius)
            err_y = y_arr - y_arr * (rayon / radius)
            return err_x, err_y, err_abs
        # UAV
        uav_err_circ_x, uav_err_circ_y, uav_err_circ_abs = erreur_cercle(uav_x_f, uav_y_f)
        # UGV
        ugv_err_circ_x, ugv_err_circ_y, ugv_err_circ_abs = erreur_cercle(ugv_x_f, ugv_y_f)

        # --- Mise à jour fenêtre 5 ---
        l_uav_err_circ_x.set_data(timestamps_uav, uav_err_circ_x)
        l_uav_err_circ_y.set_data(timestamps_uav, uav_err_circ_y)
        l_ugv_err_circ_x.set_data(timestamps_ugv, ugv_err_circ_x)
        l_ugv_err_circ_y.set_data(timestamps_ugv, ugv_err_circ_y)
        l_uav_err_circ_abs.set_data(timestamps_uav, uav_err_circ_abs)
        l_ugv_err_circ_abs.set_data(timestamps_ugv, ugv_err_circ_abs)

        for ax in axs5.flat:
            lines = ax.get_lines()
            y_arrays = [line.get_ydata() for line in lines if len(line.get_ydata()) > 0]
            if y_arrays:
                ydata = np.concatenate(y_arrays)
                lim = max(1, np.ceil(np.max(np.abs(ydata))))
                ax.set_ylim(-lim, lim)
            else:
                ax.set_ylim(-1, 1)
            ax.relim()
            ax.autoscale_view()

        fig5.canvas.draw_idle()
        
        # Référence orientation (trajectoire pour UAV, cercle pour UAV/UGV)
        # Pour UAV : tangente à la trajectoire (calculée via les ref_x, ref_y)
        ref_orientations = compute_ref_orientation(np.array(ref_x), np.array(ref_y))

        # Interpolation temporelle de l'orientation de référence à la position UAV
        from scipy.spatial import cKDTree
        ref_pts = np.column_stack((ref_x[:-1], ref_y[:-1]))
        tree = cKDTree(ref_pts)
        uav_pts = np.column_stack((uav_x_f, uav_y_f))
        _, idxs = tree.query(uav_pts)
        uav_ref_yaw = ref_orientations[idxs]
        
        ugv_pts = np.column_stack((ugv_x_f, ugv_y_f))
        _, idxs = tree.query(ugv_pts)
        ugv_ref_yaw = ref_orientations[idxs]

        # Pour le cercle : orientation vers le centre (0,0)
        uav_circle_yaw = orientation_towards_center(uav_x_f, uav_y_f)
        ugv_circle_yaw = orientation_tangent_to_circle(ugv_x_f, ugv_y_f)

        # Calcul des erreurs d'orientation
        uav_ori_err_traj = angle_difference(uav_yaws, uav_ref_yaw)
        ugv_ori_err_traj = angle_difference(ugv_yaws, ugv_ref_yaw)
        uav_ori_err_circ = angle_difference(uav_yaws, uav_circle_yaw)
        ugv_ori_err_circ = angle_difference(ugv_yaws, ugv_circle_yaw)
        ori_diff_uav_ugv = angle_difference(uav_yaws[:min_len], ugv_yaws[:min_len])

        # Mise à jour des courbes
        l_uav_ori_err_traj.set_data(timestamps_uav, uav_ori_err_traj)
        l_ugv_ori_err_taj.set_data(timestamps_ugv, ugv_ori_err_traj)
        l_uav_ori_err_cir.set_data(timestamps_uav, uav_ori_err_circ)
        l_ugv_ori_err_cir.set_data(timestamps_ugv, ugv_ori_err_circ)    
        l_ori_diff_uav_ugv.set_data(timestamps_sync, ori_diff_uav_ugv)

        # Mise à l'échelle
        for ax in axs6.flat:
            ax.set_ylim(-np.pi, np.pi)
            ax.relim()
            ax.autoscale_view()
        fig6.canvas.draw_idle()
        
        # --- Mise à jour de la table de statistiques ---
        stats_data = calculer_erreurs_stats(df_sync, ref_x, ref_y, t_start, t_end)
        stats_data.extend(stats_vitesses)

        # Supprimer l'ancien tableau
        if table_stats_holder[0]:
            table_stats_holder[0].remove()

        # Créer un nouveau tableau
        table = axs_stats.table(
            cellText=stats_data,
            colLabels=["Erreur", "Moyenne", "Variance", "Écart-type", "Min", "Max", "Médiane"],
            loc='center'
        )
        table.scale(1, 2)
        axs_stats.axis('off')
        table_stats_holder[0] = table
        fig_stats.canvas.draw_idle()
        
        # --- Sauvegarde automatique des stats ---
        output_excel = os.path.join(dossier_racine, 'stats_globales.xlsx')

        # Nom d'onglet valide pour Excel
        nom_dossier = os.path.basename(os.path.dirname(fichier_excel))
        nom_dossier_excel = "".join(c for c in nom_dossier if c not in '[]:*?/\\')[:31]

        try:
            def is_excel_valid(path):
                try:
                    load_workbook(path)
                    return True
                except Exception:
                    return False

            # Choisir le mode d'écriture
            append_mode = os.path.exists(output_excel) and is_excel_valid(output_excel)

            if append_mode:
                # On ajoute dans un fichier sain
                with pd.ExcelWriter(
                    output_excel,
                    engine='openpyxl',
                    mode='a',
                    if_sheet_exists='replace'
                ) as writer:
                    df_stats = pd.DataFrame(
                        stats_data,
                        columns=["Calcul", "Moyenne", "Variance", "Écart-type", "Min", "Max", "Médiane"]
                    )
                    df_stats.to_excel(writer, sheet_name=nom_dossier_excel, index=False)
            else:
                # On recrée le fichier
                with pd.ExcelWriter(
                    output_excel,
                    engine='openpyxl',
                    mode='w'
                ) as writer:
                    df_stats = pd.DataFrame(
                        stats_data,
                        columns=["Erreur", "Moyenne", "Variance", "Écart-type", "Min", "Max", "Médiane"]
                    )
                    df_stats.to_excel(writer, sheet_name=nom_dossier_excel, index=False)

        except Exception as e:
            print(f"Erreur lors de la sauvegarde des stats : {e}")

    slider.on_changed(update)
    update(None)

    # Fonctions de sauvegarde
    def save_graph(ax, fig, name):
        # Crée une nouvelle figure en grand format
        fig_temp = plt.figure(figsize=(16, 9))
        # Copie l'axe dans la nouvelle figure
        ax_temp = fig_temp.add_subplot(111)
        # Copie le contenu de l'axe original
        for line in ax.get_lines():
            ax_temp.plot(line.get_xdata(), line.get_ydata(), label=line.get_label(), color=line.get_color(), linestyle=line.get_linestyle())
        # Copie les labels, titre, légende
        ax_temp.set_xlabel(ax.get_xlabel())
        ax_temp.set_ylabel(ax.get_ylabel())
        ax_temp.set_title(ax.get_title())
        # Ajout du quadrillage principal et mineur
        ax_temp.grid(True, which='major', linestyle='-', linewidth=0.8, color='grey')
        ax_temp.grid(True, which='minor', linestyle=':', linewidth=0.5, color='lightgrey')
        ax_temp.minorticks_on()
        handles, labels = ax.get_legend_handles_labels()
        if handles:
            ax_temp.legend(handles, labels)
        # Mise à l'échelle automatique
        ax_temp.relim()
        ax_temp.autoscale_view()
        # Sauvegarde en grand format
        fig_temp.savefig(f'{name}_timestamp_{int(slider.val[0])}_{int(slider.val[1])}.png', dpi=300, bbox_inches='tight')
        plt.close(fig_temp)
    
    for i, btn in enumerate(save_buttons1[:6]):
        btn.on_clicked(lambda event, ax=axs1.flat[i], fig=fig1, name=f'graph1_{i}': save_graph(ax, fig, name))
    for i, btn in enumerate(save_buttons2):
        btn.on_clicked(lambda event, ax=axs2.flat[i], fig=fig2, name=f'graph2_{i}': save_graph(ax, fig, name))
    for i, btn in enumerate(save_buttons5[:6]):
        btn.on_clicked(lambda event, ax=axs5.flat[i], fig=fig5, name=f'graph5_{i}': save_graph(ax, fig, name))
    for i, btn in enumerate(save_buttons7):
        btn.on_clicked(lambda event, ax=axs7.flat[i], fig=fig7, name=f'graph7_{i}': save_graph(ax, fig, name))
    for i, btn in enumerate(save_buttons8):
        btn.on_clicked(lambda event, ax=axs8[i], fig=fig8, name=f'graph8_{i}': save_graph(ax, fig, name))

    plt.show(block=True)


def maximize_figure():
    manager = plt.get_current_fig_manager()
    try:
        manager.window.showMaximized()
    except AttributeError:
        try:
            manager.window.state('zoomed')
        except AttributeError:
            manager.full_screen_toggle()


def process_all_folders(root_dir):
    for folder_name in os.listdir(root_dir):
        folder_path = os.path.join(root_dir, folder_name)
        if not os.path.isdir(folder_path):
            continue

        print(f"Traitement du dossier : {folder_name}")
        fichiers = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
        if not fichiers:
            raise FileNotFoundError("Aucun fichier .xlsx trouvé dans le dossier courant.")
            return
        fichier_excel = fichiers[0]
        fichier_excel_path = os.path.join(folder_path, fichier_excel)
        print(f"Fichier trouvé : {fichier_excel_path}")
        # Afficher les erreurs sur la même fenêtre
        if reference_loaded:
            plot_erreurs(fichier_excel_path, ref_x, ref_y)


if __name__ == "__main__":
    dossier_racine = r"C:\Users\aurel\Bureau\s_path"  
    process_all_folders(dossier_racine)